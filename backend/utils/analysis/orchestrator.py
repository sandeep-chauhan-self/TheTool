import logging
import pandas as pd

# Import all indicators
from indicators import rsi, macd, adx, psar, ema, stochastic, cci, williams, atr, bollinger, obv, cmf

# Import enhancement modules
from utils.data.validator import DataValidator
from utils.trading.trade_validator import TradeValidator
from utils.trading.entry_calculator import EntryCalculator
from utils.trading.risk_manager import RiskManager
from utils.analysis.signal_validator import SignalValidator

logger = logging.getLogger('trading_analyzer')

# Category biases
TYPE_BIAS = {
    "trend": 1.0,
    "momentum": 1.0,
    "volatility": 0.9,
    "volume": 0.9
}

# All indicators
ALL_INDICATORS = {
    "RSI": rsi,
    "MACD": macd,
    "ADX": adx,
    "Parabolic SAR": psar,
    "EMA Crossover": ema,
    "Stochastic": stochastic,
    "CCI": cci,
    "Williams %R": williams,
    "ATR": atr,
    "Bollinger Bands": bollinger,
    "OBV": obv,
    "Chaikin Money Flow": cmf
}

def analyze_ticker(ticker, indicator_list=None, capital=None, use_demo_data=False):
    """
    REFACTORED: Now delegates to AnalysisOrchestrator for modular architecture.
    
    This function maintains backward compatibility by wrapping the new
    orchestrator-based implementation (STRUCTURAL_CAUSE_003 fix).
    
    The previous 250-line monolithic implementation has been decomposed into:
    - DataFetcher: Data acquisition and validation
    - IndicatorEngine: Technical indicator calculations
    - SignalAggregator: Vote aggregation logic
    - TradeCalculator: Entry/stop/target calculations
    - ResultFormatter: Output formatting
    
    Args:
        ticker: Stock ticker symbol
        indicator_list: List of indicator names to use (None = all)
        capital: Available capital for position sizing (optional)
        use_demo_data: Use demo data if Yahoo Finance fails (for testing)
    
    Returns:
        Dictionary with analysis results including validation
    """
    from utils.analysis_orchestrator import AnalysisOrchestrator
    
    orchestrator = AnalysisOrchestrator()
    return orchestrator.analyze(ticker, indicator_list, capital, use_demo_data)


# Legacy helper functions preserved for backward compatibility
# (moved to AnalysisOrchestrator but kept here if other modules import them)

def aggregate_votes(indicator_results):
    """
    DEPRECATED: Use SignalAggregator.aggregate_votes() instead.
    Kept for backward compatibility.
    """
    from utils.analysis_orchestrator import SignalAggregator
    return SignalAggregator.aggregate_votes(indicator_results)


def get_verdict(score):
    """
    DEPRECATED: Use SignalAggregator.get_verdict() instead.
    Kept for backward compatibility.
    """
    from utils.analysis_orchestrator import SignalAggregator
    return SignalAggregator.get_verdict(score)


def export_to_excel(result):
    """
    Export analysis result to Excel file
    
    Args:
        result: Analysis result dictionary
    
    Returns:
        Path to generated Excel file
    
    Raises:
        ValueError: If result dict is missing required keys or has invalid structure
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import os
    
    # Validate top-level required keys
    required_keys = {'ticker', 'verdict', 'score', 'entry', 'stop', 'target', 'indicators'}
    missing_keys = required_keys - set(result.keys())
    if missing_keys:
        raise ValueError(
            f"Result dictionary missing required keys: {sorted(missing_keys)}. "
            f"Required keys: {sorted(required_keys)}"
        )
    
    # Validate 'indicators' is a list
    if not isinstance(result['indicators'], list):
        raise ValueError(
            f"'indicators' must be a list, got {type(result['indicators']).__name__}"
        )
    
    # Validate each indicator has required keys
    indicator_required_keys = {'name', 'vote', 'confidence', 'category'}
    for i, indicator in enumerate(result['indicators']):
        if not isinstance(indicator, dict):
            raise ValueError(
                f"Indicator at index {i} must be a dict, got {type(indicator).__name__}"
            )
        
        missing_indicator_keys = indicator_required_keys - set(indicator.keys())
        if missing_indicator_keys:
            raise ValueError(
                f"Indicator at index {i} missing required keys: {sorted(missing_indicator_keys)}. "
                f"Required keys: {sorted(indicator_required_keys)}"
            )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.title = "Analysis Report"
    
    # Header
    ws['A1'] = "Trading Signal Analysis Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Ticker info
    ws['A3'] = "Ticker:"
    ws['B3'] = result['ticker']
    ws['A4'] = "Verdict:"
    ws['B4'] = result['verdict']
    ws['A5'] = "Score:"
    ws['B5'] = result['score']
    
    # Entry/Stop/Target
    ws['A7'] = "Entry:"
    ws['B7'] = result['entry']
    ws['A8'] = "Stop Loss:"
    ws['B8'] = result['stop']
    ws['A9'] = "Target:"
    ws['B9'] = result['target']
    
    # Indicators table
    ws['A11'] = "Indicator"
    ws['B11'] = "Vote"
    ws['C11'] = "Confidence"
    ws['D11'] = "Category"
    
    for col in ['A11', 'B11', 'C11', 'D11']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 12
    for indicator in result['indicators']:
        ws[f'A{row}'] = indicator['name']
        ws[f'B{row}'] = indicator['vote']
        ws[f'C{row}'] = indicator['confidence']
        ws[f'D{row}'] = indicator['category']
        row += 1
    
    # Save file
    output_dir = os.path.join(os.getenv('DATA_PATH', './data'), 'reports')
    os.makedirs(output_dir, exist_ok=True)
    
    filepath = os.path.join(output_dir, f"{result['ticker']}_report.xlsx")
    wb.save(filepath)
    
    return filepath
