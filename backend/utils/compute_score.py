import logging
import pandas as pd
from utils.fetch_data import fetch_ticker_data

# Import all indicators
from indicators import rsi, macd, adx, psar, ema, stochastic, cci, williams, atr, bollinger, obv, cmf

# Import enhancement modules
from utils.data_validator import DataValidator
from utils.trade_validator import TradeValidator
from utils.entry_calculator import EntryCalculator
from utils.risk_manager import RiskManager
from utils.signal_validator import SignalValidator

logger = logging.getLogger('trading_analyzer')

# Category biases
TYPE_BIAS = {
    "trend": 1.0,
    "momentum": 1.0,
    "volatility": 0.9,
    "volume": 0.9
}

# All indicators
ALL_INDICATORS = {
    "RSI": rsi,
    "MACD": macd,
    "ADX": adx,
    "Parabolic SAR": psar,
    "EMA Crossover": ema,
    "Stochastic": stochastic,
    "CCI": cci,
    "Williams %R": williams,
    "ATR": atr,
    "Bollinger Bands": bollinger,
    "OBV": obv,
    "Chaikin Money Flow": cmf
}

def analyze_ticker(ticker, indicator_list=None, capital=None, use_demo_data=False):
    """
    Analyze a ticker with technical indicators
    
    Args:
        ticker: Stock ticker symbol
        indicator_list: List of indicator names to use (None = all)
        capital: Available capital for position sizing (optional)
        use_demo_data: Use demo data if Yahoo Finance fails (for testing)
    
    Returns:
        Dictionary with analysis results including validation
    """
    try:
        # Fetch data with fallback support
        data_source = 'yahoo_finance'  # Default
        
        if use_demo_data:
            from utils.alternative_data_sources import fetch_with_fallback
            df, data_source = fetch_with_fallback(ticker, use_demo=True)
        else:
            df = fetch_ticker_data(ticker)
        
        # STEP 1: DATA VALIDATION
        # Validate OHLCV data quality
        is_valid, message, warnings = DataValidator.validate_ohlcv_data(df, ticker)
        if not is_valid:
            logger.error(f"{ticker} data validation failed: {message}")
            return {
                "ticker": ticker,
                "error": "Data validation failed",
                "validation_errors": [message] + warnings,
                "data_valid": False
            }
        
        if warnings:
            for warning in warnings:
                logger.warning(f"{ticker}: {warning}")
        
        if len(df) < 200:
            logger.warning(f"{ticker} has only {len(df)} data points, need 200")
        
        # Select indicators
        if indicator_list:
            indicators_to_use = {name: ALL_INDICATORS[name] for name in indicator_list if name in ALL_INDICATORS}
        else:
            indicators_to_use = ALL_INDICATORS
        
        # Compute all indicators
        indicator_results = []
        
        for name, indicator_module in indicators_to_use.items():
            try:
                result = indicator_module.vote_and_confidence(df)
                indicator_results.append(result)
                logger.debug(f"{ticker} - {name}: vote={result['vote']}, conf={result['confidence']}")
            except Exception as e:
                logger.error(f"Failed to compute {name} for {ticker}: {str(e)}")
        
        # Aggregate votes
        score = aggregate_votes(indicator_results)
        verdict = get_verdict(score)
        
        # STEP 2: STRATEGIC ENTRY CALCULATION
        close_price = df['Close'].iloc[-1]
        atr_value = atr.calculate_atr(df)
        
        # Determine signal for entry calculation
        signal = 'BUY' if score >= 0.4 else 'SELL' if score <= -0.4 else 'HOLD'
        
        # Use strategic entry calculator instead of current price
        entry, entry_method, _ = EntryCalculator.calculate_strategic_entry(df, close_price, signal)
        
        # Calculate stop and target using ATR
        stop = entry - (1.5 * atr_value)
        target = entry + (2 * atr_value)
        
        # STEP 3: RISK VALIDATION
        # Validate trade risk parameters
        is_risk_valid, risk_errors, risk_metrics = RiskManager.validate_trade_risk(
            entry, stop, target, signal=verdict
        )
        
        if not is_risk_valid:
            logger.warning(f"{ticker} risk validation issues: {risk_errors}")
        
        # STEP 4: SIGNAL VALIDATION
        # Collect indicator values for signal validation
        indicator_values = {}
        for ind_result in indicator_results:
            if ind_result['name'] == 'RSI':
                # Get actual RSI value from dataframe
                rsi_result = rsi.vote_and_confidence(df)
                indicator_values['RSI'] = df['RSI'].iloc[-1] if 'RSI' in df.columns else None
            elif ind_result['name'] == 'MACD':
                indicator_values['MACD'] = df['MACD'].iloc[-1] if 'MACD' in df.columns else None
                indicator_values['MACD_signal'] = df['MACD_signal'].iloc[-1] if 'MACD_signal' in df.columns else None
                indicator_values['MACD_histogram'] = df['MACD_histogram'].iloc[-1] if 'MACD_histogram' in df.columns else None
        
        # Add trend from EMA
        if score > 0.4:
            indicator_values['trend'] = 'bullish'
        elif score < -0.4:
            indicator_values['trend'] = 'bearish'
        else:
            indicator_values['trend'] = 'neutral'
        
        # Validate signal for contradictions
        signal_to_validate = 'BUY' if verdict in ['Buy', 'Strong Buy'] else 'SELL' if verdict in ['Sell', 'Strong Sell'] else None
        
        signal_valid = True
        contradictions = []
        signal_confidence = 'medium'
        alignment_score = 50
        
        if signal_to_validate:
            signal_valid, contradictions, signal_confidence, alignment_score = SignalValidator.validate_signal(
                signal_to_validate, indicator_values
            )
            
            if not signal_valid:
                logger.warning(f"{ticker} signal validation issues: {contradictions}")
        
        # Build result dictionary
        # Convert score from -1..1 range to 0..100 range for validator
        score_normalized = (score + 1) * 50  # -1 -> 0, 0 -> 50, 1 -> 100
        
        # Determine signal based on verdict
        if verdict in ['Strong Buy', 'Buy']:
            signal = 'BUY'
        elif verdict in ['Strong Sell', 'Sell']:
            signal = 'SELL'
        else:
            signal = 'HOLD'
        
        result = {
            "ticker": ticker,
            "score": round(score_normalized, 2),  # Normalized 0-100 for validator
            "score_raw": round(score, 2),  # Original -1 to 1 for display
            "verdict": verdict,
            "signal": signal,  # Required by TradeValidator
            
            # Price fields - with both naming conventions
            "entry": round(entry, 2),
            "entry_price": round(entry, 2),  # Required by TradeValidator
            "entry_method": entry_method,
            "stop": round(stop, 2),
            "stop_loss": round(stop, 2),  # Required by TradeValidator
            "target": round(target, 2),
            "target_price": round(target, 2),  # Required by TradeValidator
            "current_price": close_price,  # Required by TradeValidator
            
            "indicators": indicator_results,
            
            # Data source information
            "data_source": data_source,
            "is_demo_data": data_source == 'demo_data',
            
            # Validation results
            "data_valid": True,
            "risk_valid": is_risk_valid,
            "risk_errors": risk_errors,
            "signal_valid": signal_valid,
            "contradictions": contradictions,
            "signal_confidence": signal_confidence,
            "alignment_score": alignment_score,
            
            # Risk metrics
            "risk_metrics": {
                "shares": risk_metrics.get('shares', 0),
                "position_value": risk_metrics.get('position_value', 0),
                "risk_amount": risk_metrics.get('risk_amount', 0),
                "reward_amount": risk_metrics.get('reward_amount', 0),
                "rr_ratio": risk_metrics.get('rr_ratio', 0),
                "risk_pct": risk_metrics.get('risk_pct', 0),
                "reward_pct": risk_metrics.get('reward_pct', 0)
            } if capital else {},
            
            "success": True  # Mark as successful before validation
        }
        
        # STEP 5: TRADE VALIDATION (final check)
        if verdict in ['Buy', 'Strong Buy', 'Sell', 'Strong Sell']:
            is_valid, issues, confidence = TradeValidator.validate_analysis_result(result)
            result['trade_valid'] = is_valid
            result['trade_issues'] = issues
            result['trade_confidence'] = confidence
            
            if not is_valid:
                logger.warning(f"{ticker} trade validation failed: {issues}")
        
        return result
        
    except Exception as e:
        logger.error(f"Analysis failed for {ticker}: {str(e)}")
        raise

def aggregate_votes(indicator_results):
    """
    Aggregate indicator votes into final score
    
    Formula:
    score = SUM(vote_i * conf_i * bias_i) / SUM(conf_i * bias_i)
    
    Args:
        indicator_results: List of indicator result dictionaries
    
    Returns:
        Aggregated score between -1 and +1
    """
    numerator = 0.0
    denominator = 0.0
    
    for indicator in indicator_results:
        vote = indicator['vote']
        confidence = indicator['confidence']
        category = indicator['category']
        bias = TYPE_BIAS.get(category, 1.0)
        
        numerator += vote * confidence * bias
        denominator += confidence * bias
    
    if denominator == 0:
        return 0.0
    
    score = numerator / denominator
    return score

def get_verdict(score):
    """
    Map score to verdict text
    
    Score ranges:
    - ? 0.75: Strong Buy
    - ? 0.4: Buy
    - -0.4 to 0.4: Neutral
    - ? -0.4: Sell
    - ? -0.75: Strong Sell
    """
    if score >= 0.75:
        return "Strong Buy"
    elif score >= 0.4:
        return "Buy"
    elif score <= -0.75:
        return "Strong Sell"
    elif score <= -0.4:
        return "Sell"
    else:
        return "Neutral"

def export_to_excel(result):
    """
    Export analysis result to Excel file
    
    Args:
        result: Analysis result dictionary
    
    Returns:
        Path to generated Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import os
    
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    ws.title = "Analysis Report"
    
    # Header
    ws['A1'] = "Trading Signal Analysis Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Ticker info
    ws['A3'] = "Ticker:"
    ws['B3'] = result['ticker']
    ws['A4'] = "Verdict:"
    ws['B4'] = result['verdict']
    ws['A5'] = "Score:"
    ws['B5'] = result['score']
    
    # Entry/Stop/Target
    ws['A7'] = "Entry:"
    ws['B7'] = result['entry']
    ws['A8'] = "Stop Loss:"
    ws['B8'] = result['stop']
    ws['A9'] = "Target:"
    ws['B9'] = result['target']
    
    # Indicators table
    ws['A11'] = "Indicator"
    ws['B11'] = "Vote"
    ws['C11'] = "Confidence"
    ws['D11'] = "Category"
    
    for col in ['A11', 'B11', 'C11', 'D11']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 12
    for indicator in result['indicators']:
        ws[f'A{row}'] = indicator['name']
        ws[f'B{row}'] = indicator['vote']
        ws[f'C{row}'] = indicator['confidence']
        ws[f'D{row}'] = indicator['category']
        row += 1
    
    # Save file
    output_dir = os.path.join(os.getenv('DATA_PATH', './data'), 'reports')
    os.makedirs(output_dir, exist_ok=True)
    
    filepath = os.path.join(output_dir, f"{result['ticker']}_report.xlsx")
    wb.save(filepath)
    
    return filepath
